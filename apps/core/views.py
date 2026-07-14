from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, ProtectedError
from django.http import HttpResponse
from .models import Asset, Category, Location, PhysicalAsset, SoftwareLicense, NotebookLease, SoftwareApplication
from .forms import AssetForm, CategoryForm, LocationForm, PhysicalAssetForm, SoftwareLicenseForm, NotebookLeaseForm, SoftwareApplicationForm
from apps.users.decorators import editor_required, admin_it_required
from apps.audit.middleware import log_audit
from datetime import datetime
from decimal import Decimal, InvalidOperation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


# ── Import helper functions ──────────────────────────

def _ip_find_header(ws, required_headers):
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [str(v).strip().upper() if v else '' for v in row]
        non_empty = [v for v in vals if v]
        if len(non_empty) < 3:
            continue
        # Check that ALL required headers appear (as substring) in at least one column
        matched = all(
            any(req.upper() in col for col in vals)
            for req in required_headers
        )
        if matched:
            return row_idx, vals
    return None, None


def _ip_map_header(header_vals, field_map):
    col = {}
    for i, v in enumerate(header_vals):
        v_clean = v.strip().upper()
        if not v_clean:
            continue
        for h_name, f_name in field_map.items():
            if f_name in col:
                continue  # Already mapped
            if h_name.upper() == v_clean or h_name.upper() in v_clean:
                col[f_name] = i
                break
    return col


def _ip_get(vals, col_map, field):
    if isinstance(field, str):
        field = col_map.get(field)
    if field is None or field >= len(vals):
        return ''
    return str(vals[field]).strip() if vals[field] is not None else ''


def _ip_int(vals, col_map, field):
    raw = _ip_get(vals, col_map, field)
    if not raw:
        return None
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _ip_physical(ws, import_month="2026-07"):
    row_idx, header_vals = _ip_find_header(ws, ['ASSET NAME', 'SERIAL NUMBER'])
    if not row_idx:
        return None
    
    col_map = _ip_map_header(header_vals, {
        'ASSET NAME': 'asset_name',
        'SUB-CLASSIFICATION': 'sub_class',
        'OPERATING SYSTEM': 'operating_system',
        'BRAND': 'brand',
        'MODEL': 'model_name',
        'SERIAL NUMBER': 'serial_number',
        'OWNER': 'owner',
        'LOCATION': 'location',
        'ASSET NO': 'asset_number',
        'CONFIDENTIALITY': 'confidentiality',
        'INTEGRITY': 'integrity',
        'AVAILABILITY': 'availability',
    })
    
    results = {'success': [], 'error': []}
    
    # Clear existing physical assets for this month to avoid duplication on re-import
    PhysicalAsset.objects.filter(import_month=import_month).delete()
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=row_idx + 1, values_only=True), start=row_idx + 1):
        if not any(row):
            continue
        
        asset_name = _ip_get(row, col_map, 'asset_name')
        serial_number = _ip_get(row, col_map, 'serial_number')
        asset_number = _ip_get(row, col_map, 'asset_number') or None
        
        # Treat '-' as empty
        if serial_number in ('-', '--', 'N/A', 'n/a'):
            serial_number = ''
        if asset_number in ('-', '--', 'N/A', 'n/a'):
            asset_number = None
        
        if not serial_number and not asset_name:
            continue
        
        # Auto-generate serial if missing but asset_name exists
        if not serial_number and asset_name:
            import uuid
            serial_number = f'AUTO-{uuid.uuid4().hex[:8].upper()}'
            
        row_errors = []
        if row_errors:
            results['error'].append({
                'row': r_idx,
                'serial': serial_number or '-',
                'asset_num': asset_number or '-',
                'errors': row_errors
            })
        else:
            try:
                # Deduplicate: if this serial exists in other months, delete the old records
                PhysicalAsset.objects.filter(serial_number=serial_number).delete()

                PhysicalAsset.objects.create(
                    asset_name=asset_name,
                    sub_class=_ip_get(row, col_map, 'sub_class'),
                    operating_system=_ip_get(row, col_map, 'operating_system'),
                    brand=_ip_get(row, col_map, 'brand'),
                    model_name=_ip_get(row, col_map, 'model_name'),
                    serial_number=serial_number,
                    owner=_ip_get(row, col_map, 'owner'),
                    location=_ip_get(row, col_map, 'location'),
                    asset_number=asset_number,
                    confidentiality=_ip_int(row, col_map, 'confidentiality'),
                    integrity=_ip_int(row, col_map, 'integrity'),
                    availability=_ip_int(row, col_map, 'availability'),
                    import_month=import_month,
                )
                results['success'].append({
                    'row': r_idx,
                    'brand': _ip_get(row, col_map, 'brand') or '-',
                    'serial': serial_number,
                    'asset_num': asset_number or '-'
                })
            except Exception as e:
                results['error'].append({
                    'row': r_idx,
                    'serial': serial_number or '-',
                    'asset_num': asset_number or '-',
                    'errors': [str(e)]
                })
                
    return results


def _ip_licenses(ws, import_month="2026-07"):
    row_idx, header_vals = _ip_find_header(ws, ['NAMA SOFTWARE', 'JENIS LISENSI'])
    if not row_idx:
        return None
        
    col_map = _ip_map_header(header_vals, {
        'NAMA SOFTWARE / LISENSI': 'name',
        'JENIS LISENSI': 'license_type',
        'VENDOR / PROVIDER': 'vendor',
        'NOMOR LISENSI / SERIAL': 'license_number',
        'JUMLAH LISENSI / SEAT': 'quantity',
        'LOKASI INSTALASI / PENGGUNAAN': 'location_install',
        'STATUS': 'status',
    })
    
    results = {'success': [], 'error': []}
    
    # Clear existing software licenses for this month to avoid duplication on re-import
    SoftwareLicense.objects.filter(import_month=import_month).delete()
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=row_idx + 1, values_only=True), start=row_idx + 1):
        if not any(row):
            continue
            
        name = _ip_get(row, col_map, 'name')
        if not name:
            continue
            
        row_errors = []
        if row_errors:
            results['error'].append({
                'row': r_idx,
                'serial': _ip_get(row, col_map, 'license_number') or '-',
                'asset_num': name[:30],
                'errors': row_errors
            })
        else:
            try:
                lic_num = _ip_get(row, col_map, 'license_number')
                # Deduplicate: if name and license number exists, delete old records
                SoftwareLicense.objects.filter(name=name, license_number=lic_num).delete()

                SoftwareLicense.objects.create(
                    name=name,
                    license_type=_ip_get(row, col_map, 'license_type'),
                    vendor=_ip_get(row, col_map, 'vendor'),
                    license_number=lic_num,
                    quantity=_ip_get(row, col_map, 'quantity'),
                    location_install=_ip_get(row, col_map, 'location_install'),
                    status=_ip_get(row, col_map, 'status'),
                    import_month=import_month,
                )
                results['success'].append({
                    'row': r_idx,
                    'brand': _ip_get(row, col_map, 'vendor') or 'Software License',
                    'serial': lic_num or '-',
                    'asset_num': name[:30]
                })
            except Exception as e:
                results['error'].append({
                    'row': r_idx,
                    'serial': _ip_get(row, col_map, 'license_number') or '-',
                    'asset_num': name[:30],
                    'errors': [str(e)]
                })
            
    return results


def _ip_notebooks(ws, import_month="2026-07"):
    row_idx, header_vals = _ip_find_header(ws, ['TIPE PERANGKAT', 'NOMOR KERDUS'])
    if not row_idx:
        return None
        
    col_map = _ip_map_header(header_vals, {
        'TIPE PERANGKAT': 'device_type',
        'JENIS MS OFFICE': 'office_type',
        'EMAIL LISENSI': 'email_license',
        'NOMOR ASET': 'asset_number',
        'SERIAL NUMBER': 'serial_number',
        'NOMOR KERDUS': 'box_number',
        'USER': 'user_name',
        'TANGGAL LAYANAN AKTIVASI': 'activation_date',
        'MAC ADDRESS': 'mac_address',
        'PENYEDIA LAYANAN / PIHAK KETIGA': 'vendor_provider',
    })
    
    results = {'success': [], 'error': []}
    
    # Clear existing notebook leases for this month to avoid duplication on re-import
    NotebookLease.objects.filter(import_month=import_month).delete()
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=row_idx + 1, values_only=True), start=row_idx + 1):
        if not any(row):
            continue
            
        device_type = _ip_get(row, col_map, 'device_type')
        if not device_type:
            continue
            
        asset_number = _ip_get(row, col_map, 'asset_number') or None
        serial_number = _ip_get(row, col_map, 'serial_number') or None
        
        # Auto-generate serial if missing
        if not serial_number:
            import uuid
            serial_number = f'AUTO-NB-{uuid.uuid4().hex[:8].upper()}'
            
        row_errors = []
            
        if row_errors:
            results['error'].append({
                'row': r_idx,
                'serial': serial_number or '-',
                'asset_num': asset_number or '-',
                'errors': row_errors
            })
        else:
            try:
                # Deduplicate: if this serial exists in other months, delete the old records
                NotebookLease.objects.filter(serial_number=serial_number).delete()

                NotebookLease.objects.create(
                    device_type=device_type,
                    office_type=_ip_get(row, col_map, 'office_type'),
                    email_license=_ip_get(row, col_map, 'email_license'),
                    asset_number=asset_number,
                    serial_number=serial_number,
                    box_number=_ip_get(row, col_map, 'box_number'),
                    user_name=_ip_get(row, col_map, 'user_name'),
                    activation_date=_ip_get(row, col_map, 'activation_date'),
                    mac_address=_ip_get(row, col_map, 'mac_address'),
                    vendor_provider=_ip_get(row, col_map, 'vendor_provider'),
                    import_month=import_month,
                )
                results['success'].append({
                    'row': r_idx,
                    'brand': 'Notebook Lease',
                    'serial': serial_number or '-',
                    'asset_num': asset_number or '-'
                })
            except Exception as e:
                results['error'].append({
                    'row': r_idx,
                    'serial': serial_number or '-',
                    'asset_num': asset_number or '-',
                    'errors': [str(e)]
                })
                
    return results


def _ip_applications(ws, import_month="2026-07"):
    row_idx, header_vals = _ip_find_header(ws, ['ASSET NAME', 'IP', 'SUB-CLASSIFICATION'])
    if not row_idx:
        return None
        
    col_map = _ip_map_header(header_vals, {
        'ASSET NAME': 'asset_name',
        'IP': 'ip_address',
        'SUB-CLASSIFICATION': 'sub_class',
        'OWNER': 'owner',
        'LOCATION': 'location',
        'CONFIDENTIALITY': 'confidentiality',
        'INTEGRITY': 'integrity',
        'AVAILABILITY': 'availability',
        'VALUE': 'value',
    })
    
    results = {'success': [], 'error': []}
    
    # Clear existing applications for this month to avoid duplication on re-import
    SoftwareApplication.objects.filter(import_month=import_month).delete()
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=row_idx + 1, values_only=True), start=row_idx + 1):
        if not any(row):
            continue
            
        asset_name = _ip_get(row, col_map, 'asset_name')
        if not asset_name:
            continue
            
        row_errors = []
        if row_errors:
            results['error'].append({
                'row': r_idx,
                'serial': '-',
                'asset_num': asset_name[:30],
                'errors': row_errors
            })
        else:
            try:
                ip_addr = _ip_get(row, col_map, 'ip_address')
                # Deduplicate: if name and ip exists, delete old records
                SoftwareApplication.objects.filter(asset_name=asset_name, ip_address=ip_addr).delete()

                SoftwareApplication.objects.create(
                    asset_name=asset_name,
                    ip_address=ip_addr,
                    sub_class=_ip_get(row, col_map, 'sub_class'),
                    owner=_ip_get(row, col_map, 'owner'),
                    location=_ip_get(row, col_map, 'location'),
                    confidentiality=_ip_int(row, col_map, 'confidentiality'),
                    integrity=_ip_int(row, col_map, 'integrity'),
                    availability=_ip_int(row, col_map, 'availability'),
                    value=_ip_int(row, col_map, 'value'),
                    import_month=import_month,
                )
                results['success'].append({
                    'row': r_idx,
                    'brand': 'Software Application',
                    'serial': ip_addr or '-',
                    'asset_num': asset_name[:30]
                })
            except Exception as e:
                results['error'].append({
                    'row': r_idx,
                    'serial': _ip_get(row, col_map, 'ip_address') or '-',
                    'asset_num': asset_name[:30],
                    'errors': [str(e)]
                })
                
    return results


def _ip_classic(ws, request):
    row_idx = None
    header_map = {}
    required_headers = ['Merek', 'Nomor Seri', 'Nomor Aset', 'Kategori', 'Lokasi', 'Status']
    
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [str(v).strip() if v is not None else '' for v in row]
        if any(h in vals for h in required_headers):
            row_idx = r_idx
            for i, val in enumerate(vals):
                if val:
                    header_map[val] = i
            break
            
    if row_idx is None:
        return None
        
    missing = [h for h in ['Merek', 'Nomor Seri', 'Nomor Aset', 'Kategori', 'Lokasi', 'Status'] if h not in header_map]
    if missing:
        return None
        
    status_map = {
        'aktif': 'active',
        'perbaikan': 'maintenance',
        'dihapus': 'decommissioned',
    }
    
    cat_map = {c.name.lower(): c for c in Category.objects.all()}
    loc_map = {l.name.lower(): l for l in Location.objects.all()}
    
    results = {'success': [], 'error': []}
    created_assets = []
    
    for r_idx, row in enumerate(ws.iter_rows(min_row=row_idx + 1, values_only=True), start=row_idx + 1):
        if not any(row):
            continue
            
        def get_val(header_name):
            idx = header_map.get(header_name)
            if idx is None:
                for h_key, h_idx in header_map.items():
                    if header_name.lower() in h_key.lower():
                        idx = h_idx
                        break
            if idx is None or idx >= len(row):
                return ''
            return str(row[idx]).strip() if row[idx] is not None else ''
            
        brand = get_val('Merek')
        model = get_val('Model')
        serial = get_val('Nomor Seri')
        asset_num = get_val('Nomor Aset')
        cat_name = get_val('Kategori')
        office = get_val('Versi Office')
        license_status = get_val('Lisensi Email')
        user_pc = get_val('Pengguna')
        loc_name = get_val('Lokasi')
        acq_date_str = get_val('Tanggal Perolehan')
        price_str = get_val('Harga')
        vendor = get_val('Vendor')
        contract_end_str = get_val('Akhir Kontrak')
        status_str = get_val('Status')
        notes = get_val('Catatan')
        
        row_errors = []
        if not brand:
            row_errors.append('Merek wajib diisi')
        if not serial:
            row_errors.append('Nomor Seri wajib diisi')
        elif Asset.objects.filter(serial_number=serial).exists():
            row_errors.append(f'Nomor Seri "{serial}" sudah ada di database')
        if not asset_num:
            row_errors.append('Nomor Aset wajib diisi')
        elif Asset.objects.filter(asset_number=asset_num).exists():
            row_errors.append(f'Nomor Aset "{asset_num}" sudah ada di database')
            
        category = None
        if not cat_name:
            row_errors.append('Kategori wajib diisi')
        else:
            category = cat_map.get(cat_name.lower())
            if not category:
                valid_cats = ', '.join(cat_map.keys())
                row_errors.append(f'Kategori "{cat_name}" tidak ditemukan. Pilih: {valid_cats}')
                
        location = None
        if not loc_name:
            row_errors.append('Lokasi wajib diisi')
        else:
            location = loc_map.get(loc_name.lower())
            if not location:
                valid_locs = ', '.join(loc_map.keys())
                row_errors.append(f'Lokasi "{loc_name}" tidak ditemukan. Pilih: {valid_locs}')
                
        acq_date = None
        if not acq_date_str:
            row_errors.append('Tanggal Perolehan wajib diisi')
        else:
            try:
                acq_date = datetime.strptime(acq_date_str, '%Y-%m-%d').date()
            except ValueError:
                try:
                    acq_date = datetime.strptime(acq_date_str, '%d/%m/%Y').date()
                except ValueError:
                    row_errors.append(f'Tanggal Perolehan format salah: "{acq_date_str}" (pakai YYYY-MM-DD)')
                    
        price = None
        if price_str:
            try:
                price = Decimal(price_str.replace(',', '').replace('.', ''))
            except (InvalidOperation, ValueError):
                row_errors.append(f'Harga tidak valid: "{price_str}"')
                
        contract_end = None
        if contract_end_str:
            try:
                contract_end = datetime.strptime(contract_end_str, '%Y-%m-%d').date()
            except ValueError:
                try:
                    contract_end = datetime.strptime(contract_end_str, '%d/%m/%Y').date()
                except ValueError:
                    row_errors.append(f'Format Akhir Kontrak salah: "{contract_end_str}" (pakai YYYY-MM-DD)')
                    
        status = 'active'
        if status_str:
            status = status_map.get(status_str.lower())
            if not status:
                valid_statuses = ', '.join(status_map.keys())
                row_errors.append(f'Status "{status_str}" tidak valid. Pilih: {valid_statuses}')
                
        if row_errors:
            results['error'].append({
                'row': r_idx,
                'serial': serial or '-',
                'asset_num': asset_num or '-',
                'errors': row_errors,
            })
        else:
            try:
                asset = Asset.objects.create(
                    category=category,
                    brand=brand,
                    model_name=model,
                    serial_number=serial,
                    asset_number=asset_num,
                    office_version=office,
                    email_license_status=license_status,
                    user_pc=user_pc,
                    location=location,
                    acquisition_date=acq_date,
                    price=price,
                    vendor_name=vendor,
                    vendor_contract_end=contract_end,
                    status=status,
                    notes=notes,
                    created_by=request.user,
                )
                created_assets.append(asset)
                results['success'].append({
                    'row': r_idx,
                    'brand': brand,
                    'serial': serial,
                    'asset_num': asset_num,
                })
            except Exception as e:
                results['error'].append({
                    'row': r_idx,
                    'serial': serial or '-',
                    'asset_num': asset_num or '-',
                    'errors': [str(e)],
                })
                
    if created_assets:
        log_audit(request.user, 'import', 'Asset', None,
                  f'Import {len(created_assets)} aset dari Excel', None, request)
                  
    return results


@login_required
def asset_list(request):
    active_tab = request.GET.get('tab', 'physical')
    query = request.GET.get('q', '')

    physical_assets = PhysicalAsset.objects.all()
    software_licenses = SoftwareLicense.objects.all()
    notebook_leases = NotebookLease.objects.all()
    software_applications = SoftwareApplication.objects.all()

    if query:
        physical_assets = physical_assets.filter(
            Q(asset_name__icontains=query) |
            Q(brand__icontains=query) |
            Q(model_name__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(owner__icontains=query) |
            Q(location__icontains=query) |
            Q(asset_number__icontains=query)
        )
        software_licenses = software_licenses.filter(
            Q(name__icontains=query) |
            Q(vendor__icontains=query) |
            Q(license_number__icontains=query) |
            Q(license_type__icontains=query)
        )
        notebook_leases = notebook_leases.filter(
            Q(device_type__icontains=query) |
            Q(user_name__icontains=query) |
            Q(serial_number__icontains=query) |
            Q(asset_number__icontains=query) |
            Q(vendor_provider__icontains=query)
        )
        software_applications = software_applications.filter(
            Q(asset_name__icontains=query) |
            Q(ip_address__icontains=query) |
            Q(sub_class__icontains=query) |
            Q(owner__icontains=query) |
            Q(location__icontains=query)
        )

    context = {
        'physical_assets': physical_assets,
        'software_licenses': software_licenses,
        'notebook_leases': notebook_leases,
        'software_applications': software_applications,
        'active_tab': active_tab,
        'query': query,
    }
    return render(request, 'core/asset_list.html', context)


@login_required
def physicalasset_detail(request, pk):
    item = get_object_or_404(PhysicalAsset, pk=pk)
    return render(request, 'core/physicalasset_detail.html', {'item': item})


@login_required
def softwarelicense_detail(request, pk):
    item = get_object_or_404(SoftwareLicense, pk=pk)
    return render(request, 'core/softwarelicense_detail.html', {'item': item})


@login_required
def notebooklease_detail(request, pk):
    item = get_object_or_404(NotebookLease, pk=pk)
    return render(request, 'core/notebooklease_detail.html', {'item': item})


@login_required
def softwareapplication_detail(request, pk):
    item = get_object_or_404(SoftwareApplication, pk=pk)
    return render(request, 'core/softwareapplication_detail.html', {'item': item})


@login_required
@editor_required
def physicalasset_edit(request, pk):
    item = get_object_or_404(PhysicalAsset, pk=pk)
    if request.method == 'POST':
        form = PhysicalAssetForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            log_audit(request.user, 'update', 'PhysicalAsset', item.pk, str(item), {'asset_name': item.asset_name, 'serial_number': item.serial_number}, request)
            messages.success(request, 'Physical Asset berhasil diupdate.')
            return redirect('core:physicalasset_detail', pk=item.pk)
    else:
        form = PhysicalAssetForm(instance=item)
    return render(request, 'core/physicalasset_form.html', {'form': form, 'item': item, 'title': 'Edit Physical Asset'})


@login_required
@editor_required
def softwarelicense_edit(request, pk):
    item = get_object_or_404(SoftwareLicense, pk=pk)
    if request.method == 'POST':
        form = SoftwareLicenseForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            log_audit(request.user, 'update', 'SoftwareLicense', item.pk, str(item), {}, request)
            messages.success(request, 'Software License berhasil diupdate.')
            return redirect('core:softwarelicense_detail', pk=item.pk)
    else:
        form = SoftwareLicenseForm(instance=item)
    return render(request, 'core/softwarelicense_form.html', {'form': form, 'item': item, 'title': 'Edit Software License'})


@login_required
@editor_required
def notebooklease_edit(request, pk):
    item = get_object_or_404(NotebookLease, pk=pk)
    if request.method == 'POST':
        form = NotebookLeaseForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            form.save()
            log_audit(request.user, 'update', 'NotebookLease', item.pk, str(item), {'device_type': item.device_type, 'serial_number': item.serial_number}, request)
            messages.success(request, 'Notebook Lease berhasil diupdate.')
            return redirect('core:notebooklease_detail', pk=item.pk)
    else:
        form = NotebookLeaseForm(instance=item)
    return render(request, 'core/notebooklease_form.html', {'form': form, 'item': item, 'title': 'Edit Notebook Lease'})


@login_required
@editor_required
def softwareapplication_edit(request, pk):
    item = get_object_or_404(SoftwareApplication, pk=pk)
    if request.method == 'POST':
        form = SoftwareApplicationForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            log_audit(request.user, 'update', 'SoftwareApplication', item.pk, str(item), {}, request)
            messages.success(request, 'Software Application berhasil diupdate.')
            return redirect('core:softwareapplication_detail', pk=item.pk)
    else:
        form = SoftwareApplicationForm(instance=item)
    return render(request, 'core/softwareapplication_form.html', {'form': form, 'item': item, 'title': 'Edit Software Application'})


@login_required
def asset_detail(request, pk):
    asset = get_object_or_404(Asset.objects.select_related('category', 'location', 'created_by'), pk=pk)
    return render(request, 'core/asset_detail.html', {'asset': asset})


@login_required
@editor_required
def asset_create(request):
    if request.method == 'POST':
        form = AssetForm(request.POST)
        if form.is_valid():
            asset = form.save(commit=False)
            asset.created_by = request.user
            asset.save()
            log_audit(request.user, 'create', 'Asset', asset.pk, str(asset), {'brand': asset.brand, 'serial_number': asset.serial_number}, request)
            messages.success(request, 'Data aset berhasil disimpan.')
            return redirect('core:asset_detail', pk=asset.pk)
        messages.error(request, 'Periksa kembali form.')
    else:
        form = AssetForm()
    return render(request, 'core/asset_form.html', {'form': form, 'title': 'Tambah Aset Baru'})


@login_required
@editor_required
def asset_edit(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        form = AssetForm(request.POST, instance=asset)
        if form.is_valid():
            form.save()
            log_audit(request.user, 'update', 'Asset', asset.pk, str(asset), {'brand': asset.brand, 'serial_number': asset.serial_number}, request)
            messages.success(request, 'Data aset berhasil diupdate.')
            return redirect('core:asset_detail', pk=asset.pk)
        messages.error(request, 'Periksa kembali form.')
    else:
        form = AssetForm(instance=asset)
    return render(request, 'core/asset_form.html', {'form': form, 'title': 'Edit Aset'})


@login_required
@editor_required
def asset_delete(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        log_audit(request.user, 'delete', 'Asset', asset.pk, str(asset), {'brand': asset.brand, 'serial_number': asset.serial_number}, request)
        asset.delete()
        messages.success(request, 'Aset berhasil dihapus.')
        return redirect('core:asset_list')
    return render(request, 'core/asset_confirm_delete.html', {'asset': asset})


@login_required
def asset_export(request):
    assets = Asset.objects.select_related('category', 'location').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asset Register'

    headers = ['Merek', 'Model', 'Nomor Seri', 'Nomor Aset', 'Kategori', 'Versi Office',
               'Lisensi Email', 'Pengguna', 'Lokasi', 'Tanggal Perolehan', 'Harga',
               'Vendor', 'Akhir Kontrak Vendor', 'Status', 'Catatan']
    ws.append(headers)

    for asset in assets:
        ws.append([
            asset.brand, asset.model_name, asset.serial_number, asset.asset_number,
            asset.category.name, asset.office_version, asset.email_license_status,
            asset.user_pc, asset.location.name,
            asset.acquisition_date.isoformat() if asset.acquisition_date else '',
            float(asset.price) if asset.price else '',
            asset.vendor_name,
            asset.vendor_contract_end.isoformat() if asset.vendor_contract_end else '',
            asset.get_status_display(), asset.notes
        ])

    log_audit(request.user, 'export', 'Asset', None, 'Export Asset Register', None, request)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=asset_register.xlsx'
    wb.save(response)
    return response


@login_required
@editor_required
def asset_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template Import Aset'

    headers = ['Merek', 'Model', 'Nomor Seri', 'Nomor Aset', 'Kategori', 'Versi Office',
               'Lisensi Email', 'Pengguna', 'Lokasi', 'Tanggal Perolehan (YYYY-MM-DD)', 'Harga',
               'Vendor', 'Akhir Kontrak Vendor (YYYY-MM-DD)', 'Status', 'Catatan']

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A1A1E', end_color='1A1A1E', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D1D5'),
        right=Side(style='thin', color='D1D1D5'),
        top=Side(style='thin', color='D1D1D5'),
        bottom=Side(style='thin', color='D1D1D5'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 35

    col_widths = [15, 25, 20, 15, 15, 18, 18, 20, 25, 22, 15, 25, 22, 15, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    example = ['Lenovo', 'ThinkPad X1 Carbon Gen 12', 'LP-2026-0001', 'AST-0001',
               'Laptop', 'Microsoft 365 Apps', 'Active', 'Budi Santoso',
               'Kantor Pusat - Gedung A', '2026-01-15', 31000000,
               'PT. Anugrah Perkasa', '2029-01-15', 'Aktif', 'Contoh pengisian data']
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if col_idx in [10, 13]:
            cell.number_format = 'YYYY-MM-DD'
        if col_idx == 11:
            cell.number_format = '#,##0'

    ws.row_dimensions[2].height = 22

    # Data Validation - Kategori
    cat_names = list(Category.objects.values_list('name', flat=True))
    if cat_names:
        dv_cat = DataValidation(type='list', formula1='"' + ','.join(cat_names) + '"', allow_blank=True)
        dv_cat.error = 'Pilih kategori dari daftar yang tersedia'
        dv_cat.errorTitle = 'Kategori Tidak Valid'
        dv_cat.prompt = 'Pilih kategori aset'
        dv_cat.promptTitle = 'Kategori'
        ws.add_data_validation(dv_cat)
        dv_cat.add(f'E2:E1000')

    # Data Validation - Lokasi
    loc_names = list(Location.objects.values_list('name', flat=True))
    if loc_names:
        dv_loc = DataValidation(type='list', formula1='"' + ','.join(loc_names) + '"', allow_blank=True)
        dv_loc.error = 'Pilih lokasi dari daftar yang tersedia'
        dv_loc.errorTitle = 'Lokasi Tidak Valid'
        dv_loc.prompt = 'Pilih lokasi aset'
        dv_loc.promptTitle = 'Lokasi'
        ws.add_data_validation(dv_loc)
        dv_loc.add(f'I2:I1000')

    # Data Validation - Status
    status_list = 'Aktif,Perbaikan,Dihapus'
    dv_status = DataValidation(type='list', formula1='"' + status_list + '"', allow_blank=True)
    dv_status.error = 'Pilih status: Aktif, Perbaikan, atau Dihapus'
    dv_status.errorTitle = 'Status Tidak Valid'
    dv_status.prompt = 'Pilih status aset'
    dv_status.promptTitle = 'Status'
    ws.add_data_validation(dv_status)
    dv_status.add(f'N2:N1000')

    # Data Validation - Lisensi Email
    license_list = 'Active,Expired,Trial'
    dv_license = DataValidation(type='list', formula1='"' + license_list + '"', allow_blank=True)
    dv_license.error = 'Pilih: Active, Expired, atau Trial'
    dv_license.errorTitle = 'Lisensi Tidak Valid'
    ws.add_data_validation(dv_license)
    dv_license.add(f'G2:G1000')

    # Instruction sheet
    ws2 = wb.create_sheet(title='Petunjuk')
    instructions = [
        ['Petunjuk Pengisian Template Import Aset'],
        [''],
        ['Kolom', 'Keterangan', 'Wajib?', 'Contoh'],
        ['Merek', 'Merek/perangkat (Lenovo, Dell, HP, dll)', 'Ya', 'Lenovo'],
        ['Model', 'Nama model perangkat', 'Ya', 'ThinkPad X1 Carbon Gen 12'],
        ['Nomor Seri', 'Serial number unik per unit', 'Ya', 'LP-2026-0001'],
        ['Nomor Aset', 'Nomor inventaris aset perusahaan', 'Ya', 'AST-0001'],
        ['Kategori', 'Pilih dari dropdown: Laptop, Server, Access Point, Printer, Monitor', 'Ya', 'Laptop'],
        ['Versi Office', 'Versi Microsoft Office / Google Workspace', 'Tidak', 'Microsoft 365 Apps'],
        ['Lisensi Email', 'Status lisensi: Active, Expired, Trial', 'Tidak', 'Active'],
        ['Pengguna', 'Nama pengguna/karyawan yang memakai', 'Tidak', 'Budi Santoso'],
        ['Lokasi', 'Pilih dari dropdown lokasi yang tersedia', 'Ya', 'Kantor Pusat - Gedung A'],
        ['Tanggal Perolehan', 'Format: YYYY-MM-DD', 'Ya', '2026-01-15'],
        ['Harga', 'Harga dalam Rupiah (angka saja, tanpa Rp)', 'Tidak', '31000000'],
        ['Vendor', 'Nama vendor/supplier', 'Tidak', 'PT. Aneka Teknologi'],
        ['Akhir Kontrak Vendor', 'Format: YYYY-MM-DD', 'Tidak', '2029-01-15'],
        ['Status', 'Pilih dari dropdown: Aktif, Perbaikan, Dihapus', 'Ya', 'Aktif'],
        ['Catatan', 'Keterangan tambahan', 'Tidak', 'Unit baru'],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 3:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='E7E7E9', end_color='E7E7E9', fill_type='solid')
            elif row_idx >= 4:
                if col_idx == 3 and val == 'Ya':
                    cell.font = Font(bold=True, color='FF0000')

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 30

    log_audit(request.user, 'download_template', 'Asset', None, 'Download Template Import', None, request)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_import_aset.xlsx'
    wb.save(response)
    return response


@login_required
@editor_required
def asset_import(request):
    if request.method != 'POST':
        return render(request, 'core/asset_import.html', {'results': None})

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Pilih file Excel terlebih dahulu.')
        return render(request, 'core/asset_import.html', {'results': None})
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Format file tidak valid.')
        return render(request, 'core/asset_import.html', {'results': None})
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    except Exception:
        messages.error(request, 'Gagal membaca file Excel.')
        return render(request, 'core/asset_import.html', {'results': None})

    import_month = request.POST.get('month', '2026-07')
    sheet_results = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        res = _ip_physical(ws, import_month=import_month)
        if not res:
            res = _ip_notebooks(ws, import_month=import_month)
        if not res:
            res = _ip_licenses(ws, import_month=import_month)
        if not res:
            res = _ip_applications(ws, import_month=import_month)
        if not res:
            res = _ip_classic(ws, request)
        if res:
            sheet_results[sn] = res
    wb.close()

    # Flatten and build summary results
    total_created = 0
    total_errors = 0
    all_success = []
    all_error = []
    
    for sn, res in sheet_results.items():
        success_list = res.get('success', [])
        error_list = res.get('error', [])
        total_created += len(success_list)
        total_errors += len(error_list)
        
        for item in success_list:
            item_copy = item.copy()
            item_copy['sheet'] = sn
            all_success.append(item_copy)
            
        for item in error_list:
            item_copy = item.copy()
            item_copy['sheet'] = sn
            all_error.append(item_copy)

    total_processed = total_created + total_errors
    
    results = {
        'total': total_processed,
        'success': all_success,
        'error': all_error,
        'sheet_results': sheet_results,
    }

    if total_created or total_errors:
        messages.success(request, f'Import selesai: {total_created} berhasil, {total_errors} gagal.')
    else:
        messages.warning(request, 'Tidak ada data yang diimport. Pastikan format sheet sesuai.')
        
    return render(request, 'core/asset_import.html', {'results': results})


# ── Category CRUD ──────────────────────────

@login_required
@editor_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'core/category_list.html', {'categories': categories})


@login_required
@editor_required
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            cat = form.save()
            log_audit(request.user, 'create', 'Category', cat.pk, cat.name, None, request)
            messages.success(request, f'Kategori "{cat.name}" berhasil ditambahkan.')
            return redirect('core:category_list')
        messages.error(request, 'Periksa kembali form.')
    else:
        form = CategoryForm()
    return render(request, 'core/category_form.html', {'form': form, 'title': 'Tambah Kategori'})


@login_required
@editor_required
def category_edit(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=cat)
        if form.is_valid():
            form.save()
            log_audit(request.user, 'update', 'Category', cat.pk, cat.name, None, request)
            messages.success(request, 'Kategori berhasil diupdate.')
            return redirect('core:category_list')
        messages.error(request, 'Periksa kembali form.')
    else:
        form = CategoryForm(instance=cat)
    return render(request, 'core/category_form.html', {'form': form, 'title': 'Edit Kategori'})


@login_required
@editor_required
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        try:
            log_audit(request.user, 'delete', 'Category', cat.pk, cat.name, None, request)
            cat.delete()
            messages.success(request, 'Kategori berhasil dihapus.')
        except ProtectedError:
            messages.error(request, f'Kategori "{cat.name}" tidak dapat dihapus karena masih digunakan oleh aset.')
        return redirect('core:category_list')
    return render(request, 'core/category_confirm_delete.html', {'obj': cat, 'type': 'Kategori'})


# ── Location CRUD ──────────────────────────

@login_required
@editor_required
def location_list(request):
    locations = Location.objects.all()
    return render(request, 'core/location_list.html', {'locations': locations})


@login_required
@editor_required
def location_create(request):
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            loc = form.save()
            log_audit(request.user, 'create', 'Location', loc.pk, loc.name, None, request)
            messages.success(request, f'Lokasi "{loc.name}" berhasil ditambahkan.')
            return redirect('core:location_list')
        messages.error(request, 'Periksa kembali form.')
    else:
        form = LocationForm()
    return render(request, 'core/location_form.html', {'form': form, 'title': 'Tambah Lokasi'})


@login_required
@editor_required
def location_edit(request, pk):
    loc = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=loc)
        if form.is_valid():
            form.save()
            log_audit(request.user, 'update', 'Location', loc.pk, loc.name, None, request)
            messages.success(request, 'Lokasi berhasil diupdate.')
            return redirect('core:location_list')
        messages.error(request, 'Periksa kembali form.')
    else:
        form = LocationForm(instance=loc)
    return render(request, 'core/location_form.html', {'form': form, 'title': 'Edit Lokasi'})


@login_required
@editor_required
def location_delete(request, pk):
    loc = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        try:
            log_audit(request.user, 'delete', 'Location', loc.pk, loc.name, None, request)
            loc.delete()
            messages.success(request, 'Lokasi berhasil dihapus.')
        except ProtectedError:
            messages.error(request, f'Lokasi "{loc.name}" tidak dapat dihapus karena masih digunakan oleh aset.')
        return redirect('core:location_list')
    return render(request, 'core/category_confirm_delete.html', {'obj': loc, 'type': 'Lokasi'})


# ── Real-time Activity JSON API ──────────────────────────

from django.http import JsonResponse
from apps.audit.models import AuditLog

@login_required
def asset_activity_json(request):
    """Returns the latest audit log entries for a specific asset as JSON (for AJAX polling)."""
    model_name = request.GET.get('model', '')
    object_id = request.GET.get('id', '')

    logs = AuditLog.objects.select_related('user').filter(
        model_name=model_name,
        object_id=object_id
    ).order_by('-created_at')[:15]

    data = []
    for log in logs:
        data.append({
            'id': log.pk,
            'user': log.user.get_full_name() or log.user.email if log.user else 'System',
            'email': log.user.email if log.user else '',
            'action': log.get_action_display(),
            'action_code': log.action,
            'timestamp': log.created_at.strftime('%d %b %Y %H:%M:%S'),
            'timestamp_iso': log.created_at.isoformat(),
            'details': log.details or {},
        })

    # Check if multiple users active in last 10 minutes
    from django.utils import timezone
    from datetime import timedelta
    recent_threshold = timezone.now() - timedelta(minutes=10)
    recent_users = AuditLog.objects.filter(
        model_name=model_name,
        object_id=object_id,
        created_at__gte=recent_threshold
    ).values_list('user_id', flat=True).distinct()

    return JsonResponse({
        'logs': data,
        'multi_user_active': recent_users.count() > 1,
        'active_users_count': recent_users.count(),
    })

