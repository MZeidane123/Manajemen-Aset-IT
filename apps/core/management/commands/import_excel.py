from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from apps.core.models import (
    PhysicalAsset,
    SoftwareLicense,
    NotebookLease,
)


class Command(BaseCommand):
    help = 'Import semua sheet dari IT Asset Register Excel ke database'

    def add_arguments(self, parser):
        parser.add_argument('filepath', help='Path ke file Excel')
        parser.add_argument('--skip-existing', action='store_true', help='Skip data yang sudah ada (serial_number/asset_number duplicate)')

    def handle(self, *args, **options):
        filepath = options['filepath']
        skip_existing = options['skip_existing']

        self.stdout.write(f'Membuka file: {filepath}')
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            raise CommandError(f'Gagal membuka file: {e}')

        results = {}

        if 'Physical' in wb.sheetnames:
            results['Physical'] = self._import_physical(wb['Physical'], skip_existing)
        if 'Lisence ' in wb.sheetnames:
            results['Lisence'] = self._import_licenses(wb['Lisence '], skip_existing)
        if 'Notebook' in wb.sheetnames:
            results['Notebook'] = self._import_notebooks(wb['Notebook'], skip_existing)

        wb.close()

        self._print_summary(results)

    def _find_header_row(self, ws, required_headers):
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [str(v).strip().upper() if v else '' for v in row]
            if any(h.upper() in vals for h in required_headers):
                return row_idx, vals
        return None, None

    def _build_header_map(self, header_vals, field_map):
        col_map = {}
        for i, val in enumerate(header_vals):
            for key, field in field_map.items():
                if key.upper() == val.upper().strip():
                    col_map[field] = i
                    break
        return col_map

    def _import_physical(self, ws, skip_existing):
        created = 0
        skipped = 0
        errors = []

        required_headers = ['ASSET NAME', 'SERIAL NUMBER', 'BRAND']
        header_row, header_vals = self._find_header_row(ws, required_headers)
        if not header_row:
            return {'error': 'Header Physical tidak ditemukan (cari: ASSET NAME, SERIAL NUMBER)', 'created': 0, 'skipped': 0}

        col_map = self._build_header_map(header_vals, {
            'ASSET NAME': 'asset_name',
            'SUB-CLASSIFICATION': 'sub_class',
            'OPERATING SYSTEM': 'operating_system',
            'BRAND': 'brand',
            'MODEL': 'model_name',
            'SERIAL NUMBER': 'serial_number',
            'OWNER': 'owner',
            'LOCATION': 'location',
            'ASSET NO': 'asset_number',
            'CONFIDENTIALITY': 'confidentiality',
            'INTEGRITY': 'integrity',
            'AVAILABILITY': 'availability',
        })

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            vals = [str(v).strip() if v is not None else '' for v in row]
            if not vals[col_map.get('serial_number', 0)] and not vals[col_map.get('asset_name', 0)]:
                continue

            sn = vals[col_map.get('serial_number', 0)]
            if not sn:
                errors.append(f'Row {row_idx}: serial_number kosong, skipped')
                continue

            if skip_existing and PhysicalAsset.objects.filter(serial_number=sn).exists():
                skipped += 1
                continue

            try:
                asset_number = vals[col_map.get('asset_number', 0)] or None
                if asset_number and PhysicalAsset.objects.filter(asset_number=asset_number).exists():
                    suffix = sn.replace(' ', '')[:20] if sn else 'x'
                    asset_number = f'{asset_number}-{suffix}'
                if sn and PhysicalAsset.objects.filter(serial_number=sn).exists():
                    errors.append(f'Row {row_idx}: duplicate serial_number {sn}, skipped')
                    continue
                PhysicalAsset.objects.create(
                    asset_name=vals[col_map.get('asset_name', 0)],
                    sub_class=vals[col_map.get('sub_class', 0)],
                    operating_system=vals[col_map.get('operating_system', 0)],
                    brand=vals[col_map.get('brand', 0)],
                    model_name=vals[col_map.get('model_name', 0)],
                    serial_number=sn,
                    owner=vals[col_map.get('owner', 0)],
                    location=vals[col_map.get('location', 0)],
                    asset_number=asset_number,
                    confidentiality=_safe_int(vals, col_map.get('confidentiality')),
                    integrity=_safe_int(vals, col_map.get('integrity')),
                    availability=_safe_int(vals, col_map.get('availability')),
                )
                created += 1
            except Exception as e:
                errors.append(f'Row {row_idx}: {e}')

        return {'created': created, 'skipped': skipped, 'errors': errors[:20]}

    def _import_licenses(self, ws, skip_existing):
        created = 0
        skipped = 0
        errors = []

        required_headers = ['NAMA SOFTWARE', 'JENIS LISENSI']
        header_row, header_vals = self._find_header_row(ws, required_headers)
        if not header_row:
            return {'error': 'Header Lisence tidak ditemukan (cari: NAMA SOFTWARE / LISENSI)', 'created': 0, 'skipped': 0}

        col_map = self._build_header_map(header_vals, {
            'NAMA SOFTWARE / LISENSI': 'name',
            'JENIS LISENSI': 'license_type',
            'VENDOR / PROVIDER': 'vendor',
            'NOMOR LISENSI / SERIAL': 'license_number',
            'JUMLAH LISENSI / SEAT': 'quantity',
            'LOKASI INSTALASI / PENGGUNAAN': 'location_install',
            'STATUS': 'status',
        })

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            vals = [str(v).strip() if v is not None else '' for v in row]
            name = vals[col_map.get('name', 0)] if col_map.get('name') is not None else ''
            if not name:
                continue

            if skip_existing and SoftwareLicense.objects.filter(name=name).exists():
                skipped += 1
                continue

            try:
                SoftwareLicense.objects.create(
                    name=name,
                    license_type=vals[col_map.get('license_type', 0)] if col_map.get('license_type') is not None else '',
                    vendor=vals[col_map.get('vendor', 0)] if col_map.get('vendor') is not None else '',
                    license_number=vals[col_map.get('license_number', 0)] if col_map.get('license_number') is not None else '',
                    quantity=vals[col_map.get('quantity', 0)] if col_map.get('quantity') is not None else '',
                    location_install=vals[col_map.get('location_install', 0)] if col_map.get('location_install') is not None else '',
                    status=vals[col_map.get('status', 0)] if col_map.get('status') is not None else '',
                )
                created += 1
            except Exception as e:
                errors.append(f'Row {row_idx}: {e}')

        return {'created': created, 'skipped': skipped, 'errors': errors[:20]}

    def _import_notebooks(self, ws, skip_existing):
        created = 0
        skipped = 0
        errors = []

        required_headers = ['TIPE PERANGKAT', 'NOMOR ASET']
        header_row, header_vals = self._find_header_row(ws, required_headers)
        if not header_row:
            return {'error': 'Header Notebook tidak ditemukan (cari: TIPE PERANGKAT, NOMOR ASET)', 'created': 0, 'skipped': 0}

        col_map = self._build_header_map(header_vals, {
            'TIPE PERANGKAT': 'device_type',
            'JENIS MS OFFICE': 'office_type',
            'EMAIL LISENSI': 'email_license',
            'NOMOR ASET': 'asset_number',
            'SERIAL NUMBER': 'serial_number',
            'NOMOR KERDUS': 'box_number',
            'USER': 'user_name',
            'TANGGAL LAYANAN AKTIVASI': 'activation_date',
            'MAC ADDRESS': 'mac_address',
            'PENYEDIA LAYANAN / PIHAK KETIGA': 'vendor_provider',
        })

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            vals = [str(v).strip() if v is not None else '' for v in row]
            device_type = vals[col_map.get('device_type', 0)] if col_map.get('device_type') is not None else ''
            if not device_type:
                continue

            asset_number = vals[col_map.get('asset_number', 0)] if col_map.get('asset_number') is not None else ''
            serial_number = vals[col_map.get('serial_number', 0)] if col_map.get('serial_number') is not None else ''

            if skip_existing and asset_number and NotebookLease.objects.filter(asset_number=asset_number).exists():
                skipped += 1
                continue

            try:
                if asset_number and NotebookLease.objects.filter(asset_number=asset_number).exists():
                    asset_number = f'{asset_number}-dup'
                if serial_number and NotebookLease.objects.filter(serial_number=serial_number).exists():
                    serial_number = f'{serial_number}-dup'
                NotebookLease.objects.create(
                    device_type=device_type,
                    office_type=vals[col_map.get('office_type', 0)] if col_map.get('office_type') is not None else '',
                    email_license=vals[col_map.get('email_license', 0)] if col_map.get('email_license') is not None else '',
                    asset_number=asset_number or None,
                    serial_number=serial_number or None,
                    box_number=vals[col_map.get('box_number', 0)] if col_map.get('box_number') is not None else '',
                    user_name=vals[col_map.get('user_name', 0)] if col_map.get('user_name') is not None else '',
                    activation_date=vals[col_map.get('activation_date', 0)] if col_map.get('activation_date') is not None else '',
                    mac_address=vals[col_map.get('mac_address', 0)] if col_map.get('mac_address') is not None else '',
                    vendor_provider=vals[col_map.get('vendor_provider', 0)] if col_map.get('vendor_provider') is not None else '',
                )
                created += 1
            except Exception as e:
                errors.append(f'Row {row_idx}: {e}')

        return {'created': created, 'skipped': skipped, 'errors': errors[:20]}

    def _print_summary(self, results):
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write('HASIL IMPORT')
        self.stdout.write('=' * 60)

        total_created = 0
        total_skipped = 0
        for sheet, result in results.items():
            if 'error' in result:
                self.stdout.write(self.style.ERROR(f'\n  [{sheet}] ERROR: {result["error"]}'))
                continue

            created = result.get('created', 0)
            skipped = result.get('skipped', 0)
            errors = result.get('errors', [])
            total_created += created
            total_skipped += skipped

            self.stdout.write(f'\n  [{sheet}]')
            self.stdout.write(f'    Created: {created}')
            self.stdout.write(f'    Skipped: {skipped}')
            if errors:
                self.stdout.write(self.style.WARNING(f'    Errors: {len(errors)}'))
                for e in errors[:5]:
                    self.stdout.write(f'      - {e}')

        self.stdout.write('\n' + '-' * 60)
        self.stdout.write(f'  TOTAL Created: {total_created}')
        self.stdout.write(f'  TOTAL Skipped: {total_skipped}')
        self.stdout.write('=' * 60)


def _safe_int(vals, idx):
    if idx is None or idx >= len(vals):
        return None
    raw = vals[idx]
    if not raw:
        return None
    try:
        return int(float(raw))
    except (ValueError, TypeError, OverflowError):
        return None
